
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

wb = load_workbook('123.xlsx')
ws = wb['BookingPage']

d = ws.cell(row=4,column=1)

border = Border(top=Side(style='medium'), bottom=Side(style='thin'))
for i in range(2, 22):
    d = ws.cell(row = 4, column = i)
    if not str(d.border):
        d.border = border

border = Border(top=Side(style='thin'), bottom=Side(style='medium'))
for i in range(2, 22):
    d = ws.cell(row = 5, column = i)
    if not str(d.border):
        d.border = border

border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
for i in range(2, 16):
    d = ws.cell(row = 6, column = i)
    if not str(d.border):
        d.border = border

border = Border(top=Side(style='medium'), bottom=Side(style='thin'))
for i in range(2, 22):
    d = ws.cell(row = 8, column = i)
    if not str(d.border):
        d.border = border

ws['D9'].border = Border(left=Side(style='thin'))

border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
for j in range(10, 24):
    for i in range(2, 22):
        d = ws.cell(row = j, column = i)
        if not str(d.border):
            d.border = border

border = Border(top=Side(style='thin'), bottom=Side(style='medium'))
for i in range(2, 22):
    d = ws.cell(row = 24, column = i)
    if not str(d.border):
        d.border = border

border = Border(top=Side(style='thin'), bottom=Side(style='medium'))
for i in range(15, 20):
    d = ws.cell(row = 26, column = i)
    if not str(d.border):
        d.border = border

border = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
ws['U8'].border = border
border = Border(right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))
for i in range(9, 24):
    d = ws.cell(row = i, column = 21)
    d.border = border

ws['U24'].border = Border(right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))

wb.save('123test.xlsx')

print wb
