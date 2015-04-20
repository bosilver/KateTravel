from openpyxl import load_workbook
from excel_reader import get_excel_data
from openpyxl.styles import Border, Side

def write_booking_from(detail, activities):

    ac_list = get_excel_data('KateTravel/static/tmp.xlsx')

    wb = load_workbook('KateTravel/static/KateTravelBookingForm.xlsx')
    sheet = wb['BookingPage']
    sheet['A5'].value = detail['title']
    sheet['B5'].value = detail['last_name']
    sheet['C5'].value = detail['first_name']
    sheet['G5'].value = int(detail['phone'])
    sheet['H5'].value = detail['email']


    i = 10
    for act in activities:
        sheet.cell(row=i, column=1).value = act['time']

        (location, activity) = act['id'].split('-')
        act_detail = ac_list[int(location)]['act_list'][int(activity)]

        sheet.cell(row=i, column=4).value = '%s (%s)' % (act_detail['activity'], act_detail['time'])
        sheet.cell(row=i, column=9).value = int(act['ad_count'])
        sheet.cell(row=i, column=10).value = int (act['ch_count'])
        sheet.cell(row=i, column=12).value = float(act_detail['web_adult'])
        sheet.cell(row=i, column=13).value = 0.00 if act_detail['web_child'] == '' else float(act_detail['web_child'])
        sheet.cell(row=i, column=14).value = float(act_detail['KTL_adult'])
        sheet.cell(row=i, column=15).value = 0.00 if act_detail['KTL_child'] == '' else float(act_detail['KTL_child'])
        if act_detail['pay_deposit_ad'] == '':
            sheet.cell(row=i, column=18).value = float(act_detail['KTL_adult']) * float(act['ad_count']) + float(act_detail['KTL_child']) * float(act['ch_count'])
        else:
            if act_detail['pay_deposit_ch']:
                sheet.cell(row=i, column=17).value = float(act_detail['pay_deposit_ad']) * float(act['ad_count']) + float(act_detail['pay_deposit_ch']) * float(act['ch_count'])
            else:
                sheet.cell(row=i, column=17).value = float(act_detail['pay_deposit_ad']) * float(act['ad_count'])
        i += 1


    build_border(sheet)
    wb.save('booking_%s %s.xlsx' % (detail['first_name'], detail['last_name']))
    return 200


def build_border(ws):
    border = Border(top=Side(style='medium'), bottom=Side(style='thin'))
    for i in range(2, 22):
        d = ws.cell(row = 4, column = i)
        if not str(d.border):
            d.border = border

    border = Border(top=Side(style='thin'), bottom=Side(style='medium'))
    for i in range(2, 22):
        d = ws.cell(row = 5, column = i)
        if not str(d.border):
            d.border = border

    border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
    for i in range(2, 16):
        d = ws.cell(row = 6, column = i)
        if not str(d.border):
            d.border = border

    border = Border(top=Side(style='medium'), bottom=Side(style='thin'))
    for i in range(2, 22):
        d = ws.cell(row = 8, column = i)
        if not str(d.border):
            d.border = border

    ws['D9'].border = Border(left=Side(style='thin'))

    border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    for j in range(10, 24):
        for i in range(2, 22):
            d = ws.cell(row = j, column = i)
            if not str(d.border):
                d.border = border

    border = Border(top=Side(style='thin'), bottom=Side(style='medium'))
    for i in range(2, 22):
        d = ws.cell(row = 24, column = i)
        if not str(d.border):
            d.border = border

    border = Border(top=Side(style='thin'), bottom=Side(style='medium'))
    for i in range(15, 20):
        d = ws.cell(row = 26, column = i)
        if not str(d.border):
            d.border = border

    border = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
    ws['U8'].border = border
    border = Border(right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))
    for i in range(9, 24):
        d = ws.cell(row = i, column = 21)
        d.border = border

    ws['U24'].border = Border(right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
